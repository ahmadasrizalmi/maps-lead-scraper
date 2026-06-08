"""
Google Maps Business Scraper — Beautiful GUI with All Features
Run: streamlit run gmaps_gui.py
"""

import streamlit as st
import json
import re
import time
import urllib.parse
from datetime import datetime
from pathlib import Path
from io import BytesIO

from playwright.sync_api import sync_playwright
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from utils.email_extractor import get_website_info
from utils.proxy_rotator import ProxyRotator


# ========== HELPER FUNCTIONS ==========

def clean_text(text: str) -> str:
    if not text:
        return ''
    text = re.sub(r'[\ue000-\uf8ff]', '', text)
    text = re.sub(r'[\uf000-\ufaff]', '', text)
    text = re.sub(r'\n+', ' ', text).strip()
    text = re.sub(r'\s+', ' ', text)
    return text.strip()


def extract_listing_data(listing) -> dict:
    data = {}
    try:
        name_el = listing.locator('.qBF1Pd, .fontHeadlineSmall, [class*="qBF1Pd"]').first
        if name_el.count() > 0:
            data['name'] = name_el.inner_text().strip()

        full_text = listing.inner_text()
        lines = [l.strip() for l in full_text.split('\n') if l.strip()]

        category_keywords = [
            'Kedai', 'Kafe', 'Restoran', 'Toko', 'Klinik', 'Hotel', 'Rumah',
            'Coffee', 'Restaurant', 'Store', 'Shop', 'Clinic', 'Office',
            'Cafe', 'Bar', 'Spa', 'Salon', 'Gym', 'Studio', 'Agency',
            'Digital', 'Marketing', 'Photography', 'Catering', 'Bakery',
            'Warung', 'Mall', 'Plaza', 'Center', 'Centre', 'Aesthetic'
        ]
        for line in lines[1:4]:
            if any(kw.lower() in line.lower() for kw in category_keywords):
                cat_parts = line.split('·')
                data['category'] = cat_parts[0].strip() if cat_parts else line
                break

        for line in lines:
            rating_match = re.match(r'^(\d\.\d)$', line.strip())
            if rating_match:
                try:
                    data['rating'] = float(rating_match.group(1))
                except:
                    pass
            reviews_match = re.match(r'^\(([\d.]+)\)$', line.strip())
            if reviews_match:
                try:
                    data['reviews'] = int(reviews_match.group(1).replace('.', ''))
                except:
                    pass

        for line in lines:
            line_lower = line.lower()
            if any(kw in line_lower for kw in ['jl', 'jalan', 'street', 'no.', 'rt', 'rw', 'kec', 'kel', 'kota', 'kab', 'blok', 'gedung']):
                addr = line
                if '·' in addr:
                    parts = addr.split('·')
                    for part in parts:
                        if any(kw in part.lower() for kw in ['jl', 'jalan', 'street', 'no.', 'rt', 'rw']):
                            addr = part.strip()
                            break
                data['address'] = addr
                break

        for line in lines:
            if 'Buka' in line or 'Open' in line:
                data['status'] = 'Open'
                break
            elif 'Tutup' in line or 'Closed' in line:
                data['status'] = 'Closed'
                break
    except:
        pass
    return data


def visit_detail_page(listing, page) -> dict:
    data = {}
    try:
        name_link = listing.locator('a.hfpxzc, a[href*="/maps/place/"]').first
        if name_link.count() == 0:
            return data

        href = name_link.get_attribute('href') or ''
        if not href or '/place/' not in href:
            return data

        page.goto(href, wait_until='domcontentloaded', timeout=15000)

        try:
            page.wait_for_selector('[data-item-id*="phone"], [data-item-id*="authority"], [data-item-id*="address"]', timeout=6000)
        except:
            page.wait_for_timeout(3000)

        page.wait_for_timeout(1500)

        # Phone
        phone_els = page.locator('[data-item-id*="phone"]').all()
        for el in phone_els:
            text = el.inner_text().strip()
            if text:
                phone_match = re.search(r'(\+?\d[\d\s\-()]{6,})', text)
                if phone_match:
                    data['phone'] = phone_match.group(1).strip()
                    break

        if 'phone' not in data:
            phone_btns = page.locator('button[data-item-id*="phone"], [data-tooltip*="phone"], [data-tooltip*="telepon"]').all()
            for btn in phone_btns:
                aria = btn.get_attribute('aria-label') or btn.get_attribute('data-tooltip') or ''
                phone_match = re.search(r'(\+?\d[\d\s\-()]{6,})', aria)
                if phone_match:
                    data['phone'] = phone_match.group(1).strip()
                    break

        if 'phone' not in data:
            detail_text = page.locator('.m6QErb, .DUwDvf').first
            if detail_text.count() > 0:
                text = detail_text.inner_text()
                phone_match = re.search(r'(\+62[\d\s\-]{8,}|0[\d\s\-]{8,})', text)
                if phone_match:
                    data['phone'] = phone_match.group(1).strip()

        # Website
        website_els = page.locator('[data-item-id*="authority"], [data-item-id*="website"]').all()
        for el in website_els:
            text = el.inner_text().strip()
            if text and 'google' not in text.lower() and len(text) > 3:
                data['website'] = text
                break
            href_val = el.get_attribute('href') or ''
            if href_val and 'google.com' not in href_val and href_val.startswith('http'):
                data['website'] = href_val
                break

        if 'website' not in data:
            website_links = page.locator('a[data-item-id*="authority"], a[data-tooltip*="website"], a[data-tooltip*="Open website"]').all()
            for link in website_links:
                href_val = link.get_attribute('href') or ''
                if href_val and 'google.com' not in href_val and 'maps' not in href_val:
                    data['website'] = href_val
                    break

        # Address
        addr_els = page.locator('[data-item-id*="address"]').all()
        for el in addr_els:
            text = el.inner_text().strip()
            if text and len(text) > 5:
                data['address'] = text
                break

        # Plus code
        pluscode_el = page.locator('[data-item-id*="oloc"]').first
        if pluscode_el.count() > 0:
            text = pluscode_el.inner_text().strip()
            if text:
                data['plus_code'] = text

        # Hours
        hours_els = page.locator('[data-item-id*="hours"], .t39EBf, [aria-label*="hours"], [aria-label*="jam"]').all()
        for el in hours_els:
            aria = el.get_attribute('aria-label') or ''
            text = el.inner_text().strip()
            if aria and len(aria) > 5:
                data['hours'] = aria
                break
            elif text and len(text) > 5:
                data['hours'] = text
                break

        if 'hours' not in data:
            hours_rows = page.locator('.OqCZI tr, table tr').all()
            if hours_rows:
                hours_list = []
                for row in hours_rows[:7]:
                    try:
                        cells = row.locator('td').all()
                        if len(cells) >= 2:
                            day = cells[0].inner_text().strip()
                            time_val = cells[1].inner_text().strip()
                            if day and time_val:
                                hours_list.append(f"{day}: {time_val}")
                    except:
                        pass
                if hours_list:
                    data['hours'] = ' | '.join(hours_list)

        # Maps URL
        try:
            current_url = page.url
            if '/place/' in current_url:
                data['maps_url'] = current_url.split('&')[0]
        except:
            pass

        # Description
        about_els = page.locator('.PYvSYb, [data-attrid*="description"]').all()
        for el in about_els:
            text = el.inner_text().strip()
            if text and len(text) > 10:
                data['description'] = text[:200]
                break

    except:
        pass

    # Go back
    try:
        page.go_back(wait_until='domcontentloaded', timeout=10000)
        page.wait_for_timeout(2000)
        page.wait_for_selector('div.Nv2PK', timeout=5000)
    except:
        pass

    return data


def scrape_google_maps(query: str, max_results: int = 30, extract_emails: bool = False, progress_callback=None) -> list:
    results = []

    with sync_playwright() as p:
        browser = p.chromium.launch(
            headless=True,
            args=['--no-sandbox', '--disable-dev-shm-usage']
        )
        context = browser.new_context(
            viewport={'width': 1920, 'height': 1080},
            user_agent='Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
            locale='id-ID'
        )
        page = context.new_page()

        encoded_query = urllib.parse.quote_plus(query)
        url = f"https://www.google.com/maps/search/{encoded_query}"

        page.goto(url, wait_until='domcontentloaded', timeout=30000)
        page.wait_for_timeout(3000)

        try:
            accept_btn = page.locator('button:has-text("Accept all"), button:has-text("Terima")')
            if accept_btn.count() > 0:
                accept_btn.first.click()
                page.wait_for_timeout(1000)
        except:
            pass

        scrollable = page.locator('div[role="feed"]').first
        if scrollable.count() == 0:
            scrollable = page.locator('.m6QErb').first

        prev_count = 0
        scroll_attempts = 0

        while scroll_attempts < 30:
            listings = page.locator('div.Nv2PK').all()
            current_count = len(listings)

            if current_count >= max_results:
                break

            if current_count == prev_count:
                scroll_attempts += 1
                if scroll_attempts >= 5:
                    break
            else:
                scroll_attempts = 0

            prev_count = current_count

            if scrollable.count() > 0:
                scrollable.evaluate('el => el.scrollTop = el.scrollHeight')
            else:
                page.evaluate('window.scrollBy(0, 1000)')

            page.wait_for_timeout(1500)

        listings = page.locator('div.Nv2PK').all()
        total = min(len(listings), max_results)

        for i, listing in enumerate(listings[:max_results]):
            try:
                data = extract_listing_data(listing)
                if data and data.get('name'):
                    data = {k: clean_text(v) if isinstance(v, str) else v for k, v in data.items()}

                    detail_data = visit_detail_page(listing, page)
                    if detail_data:
                        detail_data = {k: clean_text(v) if isinstance(v, str) else v for k, v in detail_data.items()}
                        data.update(detail_data)

                    results.append(data)

                    if progress_callback:
                        progress_callback(i + 1, total, data['name'])
            except:
                pass

        browser.close()

    # Extract emails and social media
    if extract_emails and results:
        if progress_callback:
            progress_callback(0, len(results), "📧 Extracting emails & social media...")
        
        for i, result in enumerate(results):
            if result.get('website'):
                try:
                    website_info = get_website_info(result['website'])
                    
                    if website_info.get('email'):
                        result['email'] = website_info['email']
                    if website_info.get('social_media'):
                        result['social_media'] = website_info['social_media']
                        for platform, link in website_info['social_media'].items():
                            result[f'social_{platform}'] = link if isinstance(link, str) else link[0]
                    
                    if progress_callback:
                        progress_callback(i + 1, len(results), f"📧 {result['name']}")
                except:
                    pass

    return results


def create_excel(results: list, query: str) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Business Leads"

    header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='6366f1', end_color='6366f1', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    data_font = Font(name='Calibri', size=10)
    data_align = Alignment(vertical='center', wrap_text=True)
    alt_fill = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )

    columns = [
        ('No', 5), ('Business Name', 30), ('Category', 20), ('Address', 40),
        ('Phone', 18), ('Website', 30), ('Email', 28),
        ('Instagram', 25), ('Facebook', 25), ('TikTok', 25),
        ('Rating', 8), ('Reviews', 9), ('Status', 10), ('Hours', 30),
        ('Google Maps URL', 45), ('Description', 35),
        ('Outreach Status', 15), ('Notes', 25),
    ]

    # Title
    ws.merge_cells('A1:R1')
    title_cell = ws['A1']
    title_cell.value = f"📊 Business Leads — {query}"
    title_cell.font = Font(name='Calibri', bold=True, size=16, color='6366f1')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 40

    # Subtitle
    ws.merge_cells('A2:R2')
    sub_cell = ws['A2']
    sub_cell.value = f"Generated: {datetime.now().strftime('%d %B %Y, %H:%M')} • {len(results)} businesses"
    sub_cell.font = Font(name='Calibri', size=10, color='94A3B8', italic=True)
    sub_cell.alignment = Alignment(horizontal='center')
    ws.row_dimensions[2].height = 24

    # Headers
    header_row = 4
    for col_idx, (col_name, col_width) in enumerate(columns, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width
    ws.row_dimensions[header_row].height = 30

    # Data
    for idx, result in enumerate(results):
        row = header_row + 1 + idx
        social = result.get('social_media', {})

        values = [
            idx + 1, result.get('name', ''), result.get('category', ''),
            result.get('address', ''), result.get('phone', ''),
            result.get('website', ''), result.get('email', ''),
            social.get('instagram', result.get('social_instagram', '')),
            social.get('facebook', result.get('social_facebook', '')),
            social.get('tiktok', result.get('social_tiktok', '')),
            result.get('rating', ''), result.get('reviews', ''),
            result.get('status', ''), result.get('hours', ''),
            result.get('maps_url', ''), result.get('description', ''),
            '', '',
        ]

        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.font = data_font
            cell.alignment = data_align
            cell.border = thin_border
            if idx % 2 == 1:
                cell.fill = alt_fill

        if result.get('website'):
            ws.cell(row=row, column=6).font = Font(name='Calibri', size=10, color='6366f1', underline='single')
        if result.get('maps_url'):
            ws.cell(row=row, column=15).font = Font(name='Calibri', size=10, color='6366f1', underline='single')

        ws.row_dimensions[row].height = 24

    from openpyxl.worksheet.datavalidation import DataValidation
    if results:
        dv = DataValidation(
            type="list",
            formula1='"Not Contacted,Email Sent,Follow Up 1,Follow Up 2,Responded,Interested,Not Interested,Closed"',
            allow_blank=True
        )
        first_data_row = header_row + 1
        last_data_row = header_row + len(results)
        dv.add(f'Q{first_data_row}:Q{last_data_row}')
        ws.add_data_validation(dv)

    ws.freeze_panes = f'A{header_row + 1}'

    if results:
        last_col = get_column_letter(len(columns))
        ws.auto_filter.ref = f'A{header_row}:{last_col}{header_row + len(results)}'

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    summary_data = [
        ("📊 Outreach Summary", ""),
        ("", ""),
        ("Search Query", query),
        ("Total Businesses", len(results)),
        ("With Phone", sum(1 for r in results if r.get('phone'))),
        ("With Website", sum(1 for r in results if r.get('website'))),
        ("With Email", sum(1 for r in results if r.get('email'))),
        ("With Instagram", sum(1 for r in results if r.get('social_instagram'))),
        ("With Facebook", sum(1 for r in results if r.get('social_facebook'))),
        ("Generated", datetime.now().strftime('%d %B %Y, %H:%M')),
        ("", ""),
        ("📝 Outreach Tips", ""),
        ("1.", "Personalize — mention their business name"),
        ("2.", "Reference their Google reviews"),
        ("3.", "Subject line under 50 chars"),
        ("4.", "Follow up after 3-5 days"),
        ("5.", "Track with Outreach Status dropdown"),
    ]
    for row_idx, (key, val) in enumerate(summary_data, 1):
        ws2.cell(row=row_idx, column=1, value=key).font = Font(name='Calibri', bold=True, size=11)
        ws2.cell(row=row_idx, column=2, value=str(val)).font = Font(name='Calibri', size=11)
    ws2.column_dimensions['A'].width = 25
    ws2.column_dimensions['B'].width = 50
    ws2['A1'].font = Font(name='Calibri', bold=True, size=16, color='6366f1')

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


# ========== STREAMLIT UI ==========

st.set_page_config(
    page_title="Maps Lead Scraper",
    page_icon="🧲",
    layout="wide",
    initial_sidebar_state="collapsed"
)

st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&display=swap');
    
    * { font-family: 'Inter', -apple-system, BlinkMacSystemFont, sans-serif !important; }
    
    .stApp {
        background: #0a0a0f;
        color: #e2e8f0;
    }
    
    .main .block-container {
        max-width: 900px;
        padding: 2rem 1rem;
    }
    
    /* Hero Section */
    .hero {
        text-align: center;
        padding: 2rem 0 1rem;
    }
    .hero h1 {
        font-size: 2.5rem !important;
        font-weight: 800 !important;
        background: linear-gradient(135deg, #818cf8, #c084fc, #f472b6);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        margin-bottom: 0.5rem !important;
        line-height: 1.2 !important;
    }
    .hero p {
        color: #94a3b8;
        font-size: 1.05rem;
        margin: 0;
    }
    
    /* Cards */
    .card {
        background: rgba(30, 41, 59, 0.5);
        border: 1px solid rgba(99, 102, 241, 0.15);
        border-radius: 16px;
        padding: 1.5rem;
        margin: 1rem 0;
        backdrop-filter: blur(10px);
    }
    .card-header {
        display: flex;
        align-items: center;
        gap: 0.75rem;
        margin-bottom: 1rem;
    }
    .card-icon {
        width: 40px;
        height: 40px;
        border-radius: 12px;
        background: linear-gradient(135deg, #6366f1, #8b5cf6);
        display: flex;
        align-items: center;
        justify-content: center;
        font-size: 1.2rem;
    }
    .card-title {
        font-size: 1rem;
        font-weight: 700;
        color: #e2e8f0;
    }
    .card-subtitle {
        font-size: 0.8rem;
        color: #64748b;
    }
    
    /* Stat Cards */
    .stat-grid {
        display: grid;
        grid-template-columns: repeat(5, 1fr);
        gap: 1rem;
        margin: 1.5rem 0;
    }
    .stat-card {
        background: rgba(30, 41, 59, 0.6);
        border: 1px solid rgba(99, 102, 241, 0.1);
        border-radius: 14px;
        padding: 1.2rem;
        text-align: center;
    }
    .stat-icon { font-size: 1.5rem; margin-bottom: 0.5rem; }
    .stat-value {
        font-size: 1.8rem;
        font-weight: 800;
        background: linear-gradient(135deg, #818cf8, #c084fc);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
    }
    .stat-label {
        font-size: 0.75rem;
        color: #64748b;
        text-transform: uppercase;
        letter-spacing: 0.05em;
        margin-top: 0.25rem;
    }
    
    /* Query Examples */
    .examples {
        display: grid;
        grid-template-columns: repeat(2, 1fr);
        gap: 0.75rem;
        margin-top: 1rem;
    }
    .example-chip {
        background: rgba(99, 102, 241, 0.08);
        border: 1px solid rgba(99, 102, 241, 0.15);
        border-radius: 10px;
        padding: 0.6rem 1rem;
        font-size: 0.85rem;
        color: #a5b4fc;
        cursor: pointer;
        transition: all 0.2s;
    }
    .example-chip:hover {
        background: rgba(99, 102, 241, 0.15);
        border-color: rgba(99, 102, 241, 0.3);
    }
    
    /* Input styling */
    .stTextInput > div > div > input {
        background: rgba(30, 41, 59, 0.8) !important;
        border: 1px solid rgba(99, 102, 241, 0.2) !important;
        border-radius: 12px !important;
        color: #e2e8f0 !important;
        font-size: 1rem !important;
        padding: 0.75rem 1rem !important;
    }
    .stTextInput > div > div > input:focus {
        border-color: #6366f1 !important;
        box-shadow: 0 0 0 3px rgba(99, 102, 241, 0.15) !important;
    }
    .stTextInput > div > div > input::placeholder {
        color: #475569 !important;
    }
    .stTextInput label {
        color: #94a3b8 !important;
        font-weight: 600 !important;
        font-size: 0.85rem !important;
    }
    
    /* Number input */
    .stNumberInput > div > div > input {
        background: rgba(30, 41, 59, 0.8) !important;
        border: 1px solid rgba(99, 102, 241, 0.2) !important;
        border-radius: 12px !important;
        color: #e2e8f0 !important;
    }
    .stNumberInput label {
        color: #94a3b8 !important;
        font-weight: 600 !important;
        font-size: 0.85rem !important;
    }
    
    /* Checkbox */
    .stCheckbox label {
        color: #94a3b8 !important;
    }
    
    /* Buttons */
    .stButton > button {
        background: linear-gradient(135deg, #6366f1, #8b5cf6) !important;
        color: white !important;
        border: none !important;
        border-radius: 14px !important;
        padding: 0.8rem 2rem !important;
        font-weight: 700 !important;
        font-size: 1rem !important;
        letter-spacing: 0.02em !important;
        transition: all 0.3s !important;
        box-shadow: 0 4px 15px rgba(99, 102, 241, 0.3) !important;
    }
    .stButton > button:hover {
        transform: translateY(-2px) !important;
        box-shadow: 0 6px 20px rgba(99, 102, 241, 0.4) !important;
    }
    .stButton > button:active {
        transform: translateY(0) !important;
    }
    
    /* Download button */
    .stDownloadButton > button {
        background: linear-gradient(135deg, #10b981, #059669) !important;
        color: white !important;
        border: none !important;
        border-radius: 14px !important;
        padding: 0.8rem 2rem !important;
        font-weight: 700 !important;
        font-size: 1rem !important;
        box-shadow: 0 4px 15px rgba(16, 185, 129, 0.3) !important;
        width: 100% !important;
    }
    .stDownloadButton > button:hover {
        transform: translateY(-2px) !important;
        box-shadow: 0 6px 20px rgba(16, 185, 129, 0.4) !important;
    }
    
    /* Progress bar */
    .stProgress > div > div > div {
        background: linear-gradient(90deg, #6366f1, #8b5cf6, #c084fc) !important;
        border-radius: 10px !important;
    }
    .stProgress > div > div {
        background: rgba(30, 41, 59, 0.8) !important;
        border-radius: 10px !important;
        height: 10px !important;
    }
    
    /* Divider */
    hr {
        border-color: rgba(99, 102, 241, 0.15) !important;
        margin: 2rem 0 !important;
    }
    
    /* Status text */
    .status-text {
        font-size: 0.9rem;
        color: #94a3b8;
        padding: 0.5rem 0;
    }
    .status-text strong {
        color: #a5b4fc;
    }
    
    /* Footer */
    .footer {
        text-align: center;
        padding: 2rem 0 1rem;
        color: #334155;
        font-size: 0.75rem;
    }
    
    /* Sidebar hidden */
    [data-testid="stSidebar"] { display: none; }
    
    /* Columns gap fix */
    [data-testid="column"] {
        padding: 0 0.5rem !important;
    }
</style>
""", unsafe_allow_html=True)


# ========== HERO ==========
st.markdown("""
<div class="hero">
    <h1>🧲 Maps Lead Scraper</h1>
    <p>Scrape bisnis dari Google Maps → Excel siap outreach</p>
</div>
""", unsafe_allow_html=True)


# ========== SEARCH CARD ==========
st.markdown("""
<div class="card">
    <div class="card-header">
        <div class="card-icon">🔍</div>
        <div>
            <div class="card-title">Search Configuration</div>
            <div class="card-subtitle">Masukkan keyword bisnis dan lokasi target</div>
        </div>
    </div>
</div>
""", unsafe_allow_html=True)

col1, col2 = st.columns([4, 1])
with col1:
    query = st.text_input(
        "Search Query",
        placeholder="contoh: klinik kecantikan Jakarta Selatan",
        label_visibility="collapsed"
    )
with col2:
    max_results = st.number_input(
        "Max Results",
        min_value=5,
        max_value=200,
        value=20,
        step=5,
        label_visibility="collapsed"
    )

# Options
col3, col4 = st.columns(2)
with col3:
    extract_emails = st.checkbox("📧 Extract emails & social media from websites", value=False, 
                                  help="Kunjungi website bisnis untuk cari email, Instagram, Facebook, TikTok")
with col4:
    st.markdown("")  # Spacer

# Example queries
st.markdown("""
<div class="examples">
    <div class="example-chip">💆 klinik kecantikan Jakarta</div>
    <div class="example-chip">📱 digital marketing agency Surabaya</div>
    <div class="example-chip">☕ coffee shop Bandung</div>
    <div class="example-chip">🏨 hotel Bali</div>
</div>
""", unsafe_allow_html=True)

st.markdown("<br>", unsafe_allow_html=True)

# Start button
if st.button("🚀  Start Scraping", use_container_width=True):
    if not query:
        st.error("Masukkan search query dulu!")
    else:
        st.markdown("---")
        
        # Progress section
        st.markdown("""
        <div class="card">
            <div class="card-header">
                <div class="card-icon">⚡</div>
                <div>
                    <div class="card-title">Scraping in Progress</div>
                    <div class="card-subtitle">Mohon tunggu, sedang mengambil data dari Google Maps...</div>
                </div>
            </div>
        </div>
        """, unsafe_allow_html=True)
        
        progress_bar = st.progress(0)
        status_text = st.empty()
        
        results = []

        def progress_callback(current, total, name):
            if total > 0:
                progress_bar.progress(current / total)
            status_text.markdown(f"""
            <div class="status-text">
                <strong>[{current}/{total}]</strong> ✅ {name}
            </div>
            """, unsafe_allow_html=True)

        with st.spinner(""):
            try:
                results = scrape_google_maps(query, max_results, extract_emails, progress_callback)
            except Exception as e:
                st.error(f"Error: {str(e)}")

        if results:
            progress_bar.progress(1.0)
            
            # Success message
            st.markdown(f"""
            <div class="card" style="border-color: rgba(16, 185, 129, 0.3);">
                <div class="card-header">
                    <div class="card-icon" style="background: linear-gradient(135deg, #10b981, #059669);">✅</div>
                    <div>
                        <div class="card-title">Scraping Complete!</div>
                        <div class="card-subtitle">{len(results)} bisnis berhasil di-scrape dari Google Maps</div>
                    </div>
                </div>
            </div>
            """, unsafe_allow_html=True)
            
            # Stats
            phone_count = sum(1 for r in results if r.get('phone'))
            website_count = sum(1 for r in results if r.get('website'))
            email_count = sum(1 for r in results if r.get('email'))
            instagram_count = sum(1 for r in results if r.get('social_instagram'))
            
            st.markdown(f"""
            <div class="stat-grid">
                <div class="stat-card">
                    <div class="stat-icon">📊</div>
                    <div class="stat-value">{len(results)}</div>
                    <div class="stat-label">Total</div>
                </div>
                <div class="stat-card">
                    <div class="stat-icon">📞</div>
                    <div class="stat-value">{phone_count}</div>
                    <div class="stat-label">Phone</div>
                </div>
                <div class="stat-card">
                    <div class="stat-icon">🌐</div>
                    <div class="stat-value">{website_count}</div>
                    <div class="stat-label">Website</div>
                </div>
                <div class="stat-card">
                    <div class="stat-icon">📧</div>
                    <div class="stat-value">{email_count}</div>
                    <div class="stat-label">Email</div>
                </div>
                <div class="stat-card">
                    <div class="stat-icon">📱</div>
                    <div class="stat-value">{instagram_count}</div>
                    <div class="stat-label">Instagram</div>
                </div>
            </div>
            """, unsafe_allow_html=True)
            
            # Preview
            st.markdown("""
            <div class="card">
                <div class="card-header">
                    <div class="card-icon">📋</div>
                    <div>
                        <div class="card-title">Data Preview</div>
                        <div class="card-subtitle">Review data sebelum download</div>
                    </div>
                </div>
            </div>
            """, unsafe_allow_html=True)
            
            preview_data = []
            for r in results:
                social = r.get('social_media', {})
                preview_data.append({
                    "Name": r.get('name', ''),
                    "📞 Phone": r.get('phone', ''),
                    "🌐 Website": r.get('website', ''),
                    "📧 Email": r.get('email', ''),
                    "📱 Instagram": social.get('instagram', r.get('social_instagram', '')),
                    "📍 Address": (r.get('address', '')[:40] + '...') if len(r.get('address', '')) > 40 else r.get('address', ''),
                })
            st.dataframe(preview_data, use_container_width=True, hide_index=True)
            
            # Download section
            st.markdown("""
            <div class="card" style="border-color: rgba(99, 102, 241, 0.3);">
                <div class="card-header">
                    <div class="card-icon">⬇️</div>
                    <div>
                        <div class="card-title">Download Results</div>
                        <div class="card-subtitle">File Excel sudah diformat rapi untuk outreach</div>
                    </div>
                </div>
            </div>
            """, unsafe_allow_html=True)
            
            excel_data = create_excel(results, query)
            filename = f"gmaps_{re.sub(r'[^a-zA-Z0-9]', '_', query).strip('_')[:30]}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
            
            # Create CSV data
            import csv
            from io import StringIO
            csv_buffer = StringIO()
            csv_writer = csv.writer(csv_buffer)
            csv_headers = ['No', 'Business Name', 'Category', 'Address', 'Phone', 'Website', 'Email', 
                          'Instagram', 'Facebook', 'TikTok', 'Rating', 'Reviews', 'Status', 'Hours',
                          'Google Maps URL', 'Description', 'Outreach Status', 'Notes']
            csv_writer.writerow(csv_headers)
            for idx, r in enumerate(results, 1):
                social = r.get('social_media', {})
                csv_writer.writerow([
                    idx, r.get('name', ''), r.get('category', ''), r.get('address', ''),
                    r.get('phone', ''), r.get('website', ''), r.get('email', ''),
                    social.get('instagram', r.get('social_instagram', '')),
                    social.get('facebook', r.get('social_facebook', '')),
                    social.get('tiktok', r.get('social_tiktok', '')),
                    r.get('rating', ''), r.get('reviews', ''), r.get('status', ''),
                    r.get('hours', ''), r.get('maps_url', ''), r.get('description', ''),
                    '', ''
                ])
            csv_data = csv_buffer.getvalue().encode('utf-8-sig')
            
            col_dl1, col_dl2, col_dl3 = st.columns(3)
            with col_dl1:
                st.download_button(
                    label="📊  Download Excel",
                    data=excel_data,
                    file_name=filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
            with col_dl2:
                st.download_button(
                    label="📋  Download CSV",
                    data=csv_data,
                    file_name=filename.replace('.xlsx', '.csv'),
                    mime="text/csv",
                    use_container_width=True
                )
            with col_dl3:
                json_data = json.dumps(results, indent=2, ensure_ascii=False)
                st.download_button(
                    label="📄  Download JSON",
                    data=json_data,
                    file_name=filename.replace('.xlsx', '.json'),
                    mime="application/json",
                    use_container_width=True
                )
        else:
            st.warning("❌ Tidak ada hasil ditemukan. Coba query lain.")

# Footer
st.markdown("""
<div class="footer">
    Maps Lead Scraper • Built with ❤️
</div>
""", unsafe_allow_html=True)
