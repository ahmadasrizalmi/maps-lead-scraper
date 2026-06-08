#!/usr/bin/env python3
"""
Google Maps Business Scraper — Ready for Outreach
Scrape business listings from Google Maps with email, social media, and more.

Usage:
    python3 gmaps_scraper.py "restaurants in Jakarta"
    python3 gmaps_scraper.py "digital agency Surabaya" --max 50 --extract-emails
    python3 gmaps_scraper.py "klinik kecantikan Bandung" --max 100 --proxy-file proxies.txt
"""

import argparse
import json
import re
import sys
import time
import urllib.parse
from datetime import datetime
from pathlib import Path

from playwright.sync_api import sync_playwright
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Import our utilities
from utils.email_extractor import get_website_info
from utils.proxy_rotator import ProxyRotator


def clean_text(text: str) -> str:
    """Remove Google Maps icon unicode characters and clean whitespace."""
    if not text:
        return ''
    text = re.sub(r'[\ue000-\uf8ff]', '', text)
    text = re.sub(r'[\uf000-\ufaff]', '', text)
    text = re.sub(r'\n+', ' ', text).strip()
    text = re.sub(r'\s+', ' ', text)
    return text.strip()


def extract_listing_data(listing) -> dict:
    """Extract basic data from a search result listing."""
    data = {}
    
    try:
        name_el = listing.locator('.qBF1Pd, .fontHeadlineSmall, [class*="qBF1Pd"]').first
        if name_el.count() > 0:
            data['name'] = name_el.inner_text().strip()

        full_text = listing.inner_text()
        lines = [l.strip() for l in full_text.split('\n') if l.strip()]

        category_keywords = [
            'Kedai', 'Kafe', 'Restoran', 'Toko', 'Klinik', 'Hotel', 'Rumah',
            'Coffee', 'Restaurant', 'Store', 'Shop', 'Clinic', 'Office',
            'Cafe', 'Bar', 'Spa', 'Salon', 'Gym', 'Studio', 'Agency',
            'Digital', 'Marketing', 'Photography', 'Catering', 'Bakery',
            'Warung', 'Mall', 'Plaza', 'Center', 'Centre', 'Aesthetic'
        ]
        for line in lines[1:4]:
            if any(kw.lower() in line.lower() for kw in category_keywords):
                cat_parts = line.split('·')
                data['category'] = cat_parts[0].strip() if cat_parts else line
                break

        for line in lines:
            rating_match = re.match(r'^(\d\.\d)$', line.strip())
            if rating_match:
                try:
                    data['rating'] = float(rating_match.group(1))
                except:
                    pass
            reviews_match = re.match(r'^\(([\d.]+)\)$', line.strip())
            if reviews_match:
                try:
                    data['reviews'] = int(reviews_match.group(1).replace('.', ''))
                except:
                    pass

        for line in lines:
            line_lower = line.lower()
            if any(kw in line_lower for kw in ['jl', 'jalan', 'street', 'no.', 'rt', 'rw', 'kec', 'kel', 'kota', 'kab', 'blok', 'gedung']):
                addr = line
                if '·' in addr:
                    parts = addr.split('·')
                    for part in parts:
                        if any(kw in part.lower() for kw in ['jl', 'jalan', 'street', 'no.', 'rt', 'rw']):
                            addr = part.strip()
                            break
                data['address'] = addr
                break

        for line in lines:
            if 'Buka' in line or 'Open' in line:
                data['status'] = 'Open'
                break
            elif 'Tutup' in line or 'Closed' in line:
                data['status'] = 'Closed'
                break
    except:
        pass
    return data


def visit_detail_page(listing, page) -> dict:
    """Click into a listing to get detailed info (phone, website, hours)."""
    data = {}
    try:
        name_link = listing.locator('a.hfpxzc, a[href*="/maps/place/"]').first
        if name_link.count() == 0:
            return data

        href = name_link.get_attribute('href') or ''
        if not href or '/place/' not in href:
            return data

        page.goto(href, wait_until='domcontentloaded', timeout=15000)

        try:
            page.wait_for_selector('[data-item-id*="phone"], [data-item-id*="authority"], [data-item-id*="address"]', timeout=6000)
        except:
            page.wait_for_timeout(3000)

        page.wait_for_timeout(1500)

        # Phone
        phone_els = page.locator('[data-item-id*="phone"]').all()
        for el in phone_els:
            text = el.inner_text().strip()
            if text:
                phone_match = re.search(r'(\+?\d[\d\s\-()]{6,})', text)
                if phone_match:
                    data['phone'] = phone_match.group(1).strip()
                    break

        if 'phone' not in data:
            phone_btns = page.locator('button[data-item-id*="phone"], [data-tooltip*="phone"], [data-tooltip*="telepon"]').all()
            for btn in phone_btns:
                aria = btn.get_attribute('aria-label') or btn.get_attribute('data-tooltip') or ''
                phone_match = re.search(r'(\+?\d[\d\s\-()]{6,})', aria)
                if phone_match:
                    data['phone'] = phone_match.group(1).strip()
                    break

        if 'phone' not in data:
            detail_text = page.locator('.m6QErb, .DUwDvf').first
            if detail_text.count() > 0:
                text = detail_text.inner_text()
                phone_match = re.search(r'(\+62[\d\s\-]{8,}|0[\d\s\-]{8,})', text)
                if phone_match:
                    data['phone'] = phone_match.group(1).strip()

        # Website
        website_els = page.locator('[data-item-id*="authority"], [data-item-id*="website"]').all()
        for el in website_els:
            text = el.inner_text().strip()
            if text and 'google' not in text.lower() and len(text) > 3:
                data['website'] = text
                break
            href_val = el.get_attribute('href') or ''
            if href_val and 'google.com' not in href_val and href_val.startswith('http'):
                data['website'] = href_val
                break

        if 'website' not in data:
            website_links = page.locator('a[data-item-id*="authority"], a[data-tooltip*="website"], a[data-tooltip*="Open website"]').all()
            for link in website_links:
                href_val = link.get_attribute('href') or ''
                if href_val and 'google.com' not in href_val and 'maps' not in href_val:
                    data['website'] = href_val
                    break

        # Address
        addr_els = page.locator('[data-item-id*="address"]').all()
        for el in addr_els:
            text = el.inner_text().strip()
            if text and len(text) > 5:
                data['address'] = text
                break

        # Plus code
        pluscode_el = page.locator('[data-item-id*="oloc"]').first
        if pluscode_el.count() > 0:
            text = pluscode_el.inner_text().strip()
            if text:
                data['plus_code'] = text

        # Hours
        hours_els = page.locator('[data-item-id*="hours"], .t39EBf, [aria-label*="hours"], [aria-label*="jam"]').all()
        for el in hours_els:
            aria = el.get_attribute('aria-label') or ''
            text = el.inner_text().strip()
            if aria and len(aria) > 5:
                data['hours'] = aria
                break
            elif text and len(text) > 5:
                data['hours'] = text
                break

        if 'hours' not in data:
            hours_rows = page.locator('.OqCZI tr, table tr').all()
            if hours_rows:
                hours_list = []
                for row in hours_rows[:7]:
                    try:
                        cells = row.locator('td').all()
                        if len(cells) >= 2:
                            day = cells[0].inner_text().strip()
                            time_val = cells[1].inner_text().strip()
                            if day and time_val:
                                hours_list.append(f"{day}: {time_val}")
                    except:
                        pass
                if hours_list:
                    data['hours'] = ' | '.join(hours_list)

        # Maps URL
        try:
            current_url = page.url
            if '/place/' in current_url:
                data['maps_url'] = current_url.split('&')[0]
        except:
            pass

        # Description
        about_els = page.locator('.PYvSYb, [data-attrid*="description"]').all()
        for el in about_els:
            text = el.inner_text().strip()
            if text and len(text) > 10:
                data['description'] = text[:200]
                break

    except:
        pass

    # Go back
    try:
        page.go_back(wait_until='domcontentloaded', timeout=10000)
        page.wait_for_timeout(2000)
        page.wait_for_selector('div.Nv2PK', timeout=5000)
    except:
        pass

    return data


def scrape_google_maps(
    query: str,
    max_results: int = 30,
    headless: bool = True,
    extract_emails: bool = False,
    proxy_rotator: ProxyRotator = None,
    progress_callback=None
) -> list[dict]:
    """Scrape business listings from Google Maps search results."""
    
    results = []
    
    with sync_playwright() as p:
        # Configure browser
        launch_args = ['--no-sandbox', '--disable-dev-shm-usage']
        
        browser_kwargs = {
            'headless': headless,
            'args': launch_args,
        }
        
        # Add proxy if available
        if proxy_rotator and proxy_rotator.has_proxies:
            proxy_config = proxy_rotator.get_playwright_proxy()
            if proxy_config:
                browser_kwargs['proxy'] = proxy_config
                print(f"🔄 Using proxy: {proxy_config['server']}")
        
        browser = p.chromium.launch(**browser_kwargs)
        
        context = browser.new_context(
            viewport={'width': 1920, 'height': 1080},
            user_agent='Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
            locale='id-ID'
        )
        page = context.new_page()
        
        # Navigate to Google Maps search
        encoded_query = urllib.parse.quote_plus(query)
        url = f"https://www.google.com/maps/search/{encoded_query}"
        
        print(f"🔍 Searching: {query}")
        print(f"📡 URL: {url}")
        
        page.goto(url, wait_until='domcontentloaded', timeout=30000)
        page.wait_for_timeout(3000)
        
        # Accept cookies if prompted
        try:
            accept_btn = page.locator('button:has-text("Accept all"), button:has-text("Terima")')
            if accept_btn.count() > 0:
                accept_btn.first.click()
                page.wait_for_timeout(1000)
        except:
            pass
        
        # Find the scrollable results panel
        scrollable = page.locator('div[role="feed"]').first
        if scrollable.count() == 0:
            scrollable = page.locator('.m6QErb').first
        
        # Scroll to load more results
        print(f"📜 Loading results (max: {max_results})...")
        prev_count = 0
        scroll_attempts = 0
        max_scroll_attempts = 30
        
        while scroll_attempts < max_scroll_attempts:
            listings = page.locator('div.Nv2PK').all()
            current_count = len(listings)
            
            if current_count >= max_results:
                print(f"   ✅ Reached {current_count} results")
                break
            
            if current_count == prev_count:
                scroll_attempts += 1
                if scroll_attempts >= 5:
                    print(f"   ℹ️  No more results loading ({current_count} total)")
                    break
            else:
                scroll_attempts = 0
                print(f"   📊 Loaded {current_count} results...")
            
            prev_count = current_count
            
            if scrollable.count() > 0:
                scrollable.evaluate('el => el.scrollTop = el.scrollHeight')
            else:
                page.evaluate('window.scrollBy(0, 1000)')
            
            page.wait_for_timeout(1500)
        
        # Extract business data
        listings = page.locator('div.Nv2PK').all()
        total = min(len(listings), max_results)
        print(f"\n📋 Extracting data from {total} listings...")
        
        for i, listing in enumerate(listings[:max_results]):
            try:
                data = extract_listing_data(listing)
                if data and data.get('name'):
                    data = {k: clean_text(v) if isinstance(v, str) else v for k, v in data.items()}
                    
                    detail_data = visit_detail_page(listing, page)
                    if detail_data:
                        detail_data = {k: clean_text(v) if isinstance(v, str) else v for k, v in detail_data.items()}
                        data.update(detail_data)
                    
                    results.append(data)
                    print(f"   [{i+1}/{total}] ✅ {data['name']}")
                    
                    if progress_callback:
                        progress_callback(i + 1, total, data['name'])
                        
            except Exception as e:
                print(f"   [{i+1}] ❌ Error: {str(e)[:60]}")
                continue
        
        browser.close()
    
    # Extract emails and social media from websites
    if extract_emails and results:
        print(f"\n📧 Extracting emails & social media from {len(results)} websites...")
        for i, result in enumerate(results):
            if result.get('website'):
                try:
                    print(f"   [{i+1}/{len(results)}] 🔍 {result['website']}...", end='', flush=True)
                    website_info = get_website_info(result['website'])
                    
                    if website_info.get('email'):
                        result['email'] = website_info['email']
                    if website_info.get('emails'):
                        result['all_emails'] = website_info['emails']
                    if website_info.get('social_media'):
                        result['social_media'] = website_info['social_media']
                        for platform, link in website_info['social_media'].items():
                            result[f'social_{platform}'] = link if isinstance(link, str) else link[0]
                    if website_info.get('website_phones'):
                        result['website_phones'] = website_info['website_phones']
                    
                    email_found = '✅' if website_info.get('email') else '❌'
                    social_count = len(website_info.get('social_media', {}))
                    print(f" {email_found} email, {social_count} socials")
                    
                except Exception as e:
                    print(f" ❌ {str(e)[:40]}")
    
    return results


def save_to_excel(results: list[dict], output_file: str, query: str):
    """Save results to a formatted Excel file ready for outreach."""
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Business Leads"
    
    # Styles
    header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='6366f1', end_color='6366f1', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    data_font = Font(name='Calibri', size=10)
    data_align = Alignment(vertical='center', wrap_text=True)
    alt_fill = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )
    
    # Columns - simplified for outreach
    columns = [
        ('No', 5),
        ('Business Name', 30),
        ('Phone', 18),
        ('Email', 28),
        ('Address', 40),
        ('Website', 30),
        ('WhatsApp', 25),
        ('Google Maps URL', 45),
    ]
    
    # Title
    ws.merge_cells('A1:H1')
    title_cell = ws['A1']
    title_cell.value = f"📊 Business Leads — {query}"
    title_cell.font = Font(name='Calibri', bold=True, size=16, color='6366f1')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 40
    
    # Subtitle
    ws.merge_cells('A2:H2')
    sub_cell = ws['A2']
    sub_cell.value = f"Generated: {datetime.now().strftime('%d %B %Y, %H:%M')} • {len(results)} businesses"
    sub_cell.font = Font(name='Calibri', size=10, color='94A3B8', italic=True)
    sub_cell.alignment = Alignment(horizontal='center')
    ws.row_dimensions[2].height = 24
    
    # Headers
    header_row = 4
    for col_idx, (col_name, col_width) in enumerate(columns, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width
    ws.row_dimensions[header_row].height = 30
    
    # Data
    for idx, result in enumerate(results):
        row = header_row + 1 + idx
        social = result.get('social_media', {})
        
        # Get WhatsApp from social_media dict
        social = result.get('social_media', {})
        whatsapp = social.get('whatsapp', result.get('social_whatsapp', ''))
        
        values = [
            idx + 1,
            result.get('name', ''),
            result.get('phone', ''),
            result.get('email', ''),
            result.get('address', ''),
            result.get('website', ''),
            whatsapp,
            result.get('maps_url', ''),
        ]
        
        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.font = data_font
            cell.alignment = data_align
            cell.border = thin_border
            if idx % 2 == 1:
                cell.fill = alt_fill
        
        # Make URLs clickable
        # Make URLs clickable
        if result.get('website'):
            ws.cell(row=row, column=6).font = Font(name='Calibri', size=10, color='6366f1', underline='single')
        if result.get('maps_url'):
            ws.cell(row=row, column=8).font = Font(name='Calibri', size=10, color='6366f1', underline='single')
        
        ws.row_dimensions[row].height = 24
    
    # Dropdown for Outreach Status
    from openpyxl.worksheet.datavalidation import DataValidation
    if results:
        dv = DataValidation(
            type="list",
            formula1='"Not Contacted,Email Sent,Follow Up 1,Follow Up 2,Responded,Interested,Not Interested,Closed"',
            allow_blank=True
        )
        first_data_row = header_row + 1
        last_data_row = header_row + len(results)
        dv.add(f'H{first_data_row}:H{last_data_row}')
        ws.add_data_validation(dv)
    
    # Freeze panes
    ws.freeze_panes = f'A{header_row + 1}'
    
    # Auto filter
    if results:
        last_col = get_column_letter(len(columns))
        ws.auto_filter.ref = f'A{header_row}:{last_col}{header_row + len(results)}'
    
    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    summary_data = [
        ("📊 Outreach Summary", ""),
        ("", ""),
        ("Search Query", query),
        ("Total Businesses", len(results)),
        ("With Phone", sum(1 for r in results if r.get('phone'))),
        ("With Website", sum(1 for r in results if r.get('website'))),
        ("With Email", sum(1 for r in results if r.get('email'))),
        ("With Instagram", sum(1 for r in results if r.get('social_instagram') or r.get('social_media', {}).get('instagram'))),
        ("With Facebook", sum(1 for r in results if r.get('social_facebook') or r.get('social_media', {}).get('facebook'))),
        ("Generated", datetime.now().strftime('%d %B %Y, %H:%M')),
        ("", ""),
        ("📝 Outreach Tips", ""),
        ("1.", "Personalize — mention their business name"),
        ("2.", "Reference their Google reviews"),
        ("3.", "Subject line under 50 chars"),
        ("4.", "Follow up after 3-5 days"),
        ("5.", "Track with Outreach Status dropdown"),
    ]
    for row_idx, (key, val) in enumerate(summary_data, 1):
        ws2.cell(row=row_idx, column=1, value=key).font = Font(name='Calibri', bold=True, size=11)
        ws2.cell(row=row_idx, column=2, value=str(val)).font = Font(name='Calibri', size=11)
    ws2.column_dimensions['A'].width = 25
    ws2.column_dimensions['B'].width = 50
    ws2['A1'].font = Font(name='Calibri', bold=True, size=16, color='6366f1')
    
    # Save
    wb.save(output_file)
    print(f"\n💾 Saved to: {output_file}")
    print(f"   📊 {len(results)} businesses exported")
    print(f"   📞 {sum(1 for r in results if r.get('phone'))} with phone numbers")
    print(f"   🌐 {sum(1 for r in results if r.get('website'))} with websites")
    print(f"   📧 {sum(1 for r in results if r.get('email'))} with emails")
    print(f"   📱 {sum(1 for r in results if r.get('social_instagram') or r.get('social_media', {}).get('instagram'))} with Instagram")


def main():
    parser = argparse.ArgumentParser(
        description='Google Maps Business Scraper — Ready for Outreach',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python3 gmaps_scraper.py "restaurants in Jakarta"
  python3 gmaps_scraper.py "digital marketing agency Surabaya" --max 50 --extract-emails
  python3 gmaps_scraper.py "klinik kecantikan Bandung" --max 100 --proxy-file proxies.txt
  python3 gmaps_scraper.py "hotel Bali" --max 200 --visible
        """
    )
    parser.add_argument('query', help='Search query for Google Maps')
    parser.add_argument('--max', type=int, default=30, help='Maximum results to scrape (default: 30)')
    parser.add_argument('--output', '-o', help='Output Excel file (default: auto-generated)')
    parser.add_argument('--visible', action='store_true', help='Show browser window (non-headless)')
    parser.add_argument('--extract-emails', action='store_true', help='Visit websites to extract emails & social media')
    parser.add_argument('--proxy-file', help='File with proxy list (one per line)')
    parser.add_argument('--proxies', help='Comma-separated proxy list')
    parser.add_argument('--format', choices=['xlsx', 'csv', 'both'], default='xlsx', help='Output format: xlsx, csv, or both (default: xlsx)')
    
    args = parser.parse_args()
    
    # Generate output filename
    if not args.output:
        safe_query = re.sub(r'[^a-zA-Z0-9\s]', '', args.query).strip().replace(' ', '_')[:50]
        timestamp = datetime.now().strftime('%Y%m%d_%H%M')
        args.output = f"gmaps_{safe_query}_{timestamp}.xlsx"
    
    # Setup proxy rotator
    proxy_rotator = None
    if args.proxy_file:
        proxy_rotator = ProxyRotator.from_file(args.proxy_file)
        print(f"🔄 Loaded {proxy_rotator.count} proxies from {args.proxy_file}")
    elif args.proxies:
        proxy_rotator = ProxyRotator.from_string(args.proxies)
        print(f"🔄 Loaded {proxy_rotator.count} proxies")
    
    print("=" * 60)
    print("🗺️  Google Maps Business Scraper")
    print("=" * 60)
    print(f"🔍 Query: {args.query}")
    print(f"📊 Max results: {args.max}")
    print(f"📁 Output: {args.output}")
    print(f"📧 Extract emails: {'Yes' if args.extract_emails else 'No'}")
    print(f"🔄 Proxies: {proxy_rotator.count if proxy_rotator else 0}")
    print(f"🖥️  Mode: {'visible' if args.visible else 'headless'}")
    print("=" * 60)
    print()
    
    # Scrape
    results = scrape_google_maps(
        query=args.query,
        max_results=args.max,
        headless=not args.visible,
        extract_emails=args.extract_emails,
        proxy_rotator=proxy_rotator,
    )
    
    if not results:
        print("\n❌ No results found. Try a different search query.")
        sys.exit(1)
    
    # Save based on format
    if args.format in ('xlsx', 'both'):
        save_to_excel(results, args.output, args.query)
    
    if args.format == 'csv':
        csv_file = args.output.replace('.xlsx', '.csv')
        save_to_csv(results, csv_file)
    elif args.format == 'both':
        csv_file = args.output.replace('.xlsx', '.csv')
        save_to_csv(results, csv_file)
    
    # Always save JSON backup
    json_file = args.output.replace('.xlsx', '.json')
    with open(json_file, 'w', encoding='utf-8') as f:
        json.dump(results, f, indent=2, ensure_ascii=False)
    print(f"📄 Raw JSON backup: {json_file}")
    
    print("\n✅ Done! Open the Excel file to start your outreach campaign.")
    print(f"   💡 Tip: Use 'Outreach Status' dropdown to track your progress")
    if args.format in ('csv', 'both'):
        print(f"   📊 Google Sheets: Upload the CSV file to Google Sheets (File → Import)")


def save_to_csv(results: list[dict], output_file: str):
    """Save results to CSV file (can be imported to Google Sheets)."""
    import csv
    
    headers = [
        'No', 'Business Name', 'Phone', 'Email', 'Address',
        'Website', 'WhatsApp', 'Google Maps URL'
    ]
    
    with open(output_file, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        
        for idx, result in enumerate(results, 1):
            social = result.get('social_media', {})
            whatsapp = social.get('whatsapp', result.get('social_whatsapp', ''))
            writer.writerow([
                idx,
                result.get('name', ''),
                result.get('phone', ''),
                result.get('email', ''),
                result.get('address', ''),
                result.get('website', ''),
                whatsapp,
                result.get('maps_url', ''),
            ])
    
    print(f"📄 CSV saved to: {output_file}")


if __name__ == '__main__':
    main()
